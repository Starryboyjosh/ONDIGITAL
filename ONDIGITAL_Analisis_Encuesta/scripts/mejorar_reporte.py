#!/usr/bin/env python3
"""Regenera el libro ejecutivo y el explorador local a partir de la base limpia."""

from __future__ import annotations

import csv
import json
import re
from collections import Counter
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

try:
    from openpyxl.drawing.image import Image as XLImage
except ImportError:  # pragma: no cover - openpyxl is required for this deliverable
    XLImage = None


ROOT = Path(__file__).resolve().parents[1]
DATA_PATH = ROOT / "datos_limpios_y_enriquecidos.csv"
METRICS_PATH = ROOT / "resumen_metricas.json"
OUTPUT_XLSX = ROOT / "ONDIGITAL_Analisis_Completo.xlsx"
OUTPUT_EXPLORER = ROOT / "explorador_respuestas.html"

NAVY = "071426"
NAVY_2 = "0D1D33"
BLUE = "3B82F6"
MINT = "00E5B0"
GREEN = "49D18E"
AMBER = "F9B84A"
CORAL = "FF6B75"
INK = "0B1A2E"
MUTED = "5D6D81"
LINE = "D5DFEB"
PALE_BLUE = "EAF3FF"
PALE_MINT = "E6FBF5"
PALE_AMBER = "FFF5DD"
PALE_CORAL = "FFE9EC"
WHITE = "FFFFFF"


def load_rows() -> tuple[list[dict[str, str]], dict]:
    with DATA_PATH.open(encoding="utf-8-sig", newline="") as source:
        rows = list(csv.DictReader(source))
    with METRICS_PATH.open(encoding="utf-8") as source:
        metrics = json.load(source)
    return rows, metrics


def parse_number(value: str, default: float = 0.0) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def parse_open_answer(value: str) -> tuple[str, str, str]:
    match = re.match(r"Mejoraría (.+?) porque (.+?) para (.+?)\.?$", value or "")
    if not match:
        return "Sin clasificar", "Sin clasificar", "Sin clasificar"
    return tuple(part.strip() for part in match.groups())


def split_values(value: str) -> list[str]:
    return [part.strip() for part in (value or "").split(";") if part.strip()]


def enrich_rows(rows: list[dict[str, str]]) -> list[dict]:
    headers = list(rows[0])
    output = []
    for index, row in enumerate(rows, 1):
        process, problem, goal = parse_open_answer(row[headers[10]])
        storage = split_values(row[headers[3]])
        output.append(
            {
                "id": index,
                "timestamp": row[headers[0]],
                "madurez": row[headers[1]],
                "horas": row[headers[2]],
                "almacenamiento": storage,
                "dolor": int(parse_number(row[headers[4]])),
                "disposicion": row[headers[5]],
                "freno": row[headers[6]],
                "externo": row[headers[7]],
                "aprendizaje": row[headers[8]],
                "factores": split_values(row[headers[9]]),
                "respuesta": row[headers[10]],
                "proceso": process,
                "problema": problem,
                "objetivo": goal,
                "score_madurez": int(parse_number(row[headers[11]])),
                "horas_estimadas": parse_number(row[headers[12]]),
                "score_dolor": int(parse_number(row[headers[13]])),
                "score_disposicion": int(parse_number(row[headers[14]])),
                "score_externo": int(parse_number(row[headers[15]])),
                "score_aprendizaje": int(parse_number(row[headers[16]])),
                "oportunidad": parse_number(row[headers[17]]),
                "nivel": row[headers[18]],
                "segmento": row[headers[19]],
            }
        )
    return output


def pct(value: int | float, total: int) -> float:
    return value / total if total else 0


def summary(rows: list[dict]) -> dict:
    total = len(rows)
    manual = sum(
        1
        for row in rows
        if any("libreta" in item.lower() or "excel" in item.lower() or "hoja" in item.lower() for item in row["almacenamiento"])
    )
    high_hours = sum(row["horas_estimadas"] >= 7.5 for row in rows)
    low_tech = sum(row["score_madurez"] <= 2 for row in rows)
    pain = sum(row["score_dolor"] >= 4 for row in rows)
    willing = sum(row["score_disposicion"] >= 3 for row in rows)
    external = sum(row["score_externo"] >= 3 for row in rows)
    learning = sum(row["score_aprendizaje"] >= 3 for row in rows)
    processes = Counter(row["proceso"] for row in rows)
    goals = Counter(row["objetivo"] for row in rows)
    barriers = Counter(row["freno"] for row in rows)
    return {
        "total": total,
        "manual": manual,
        "high_hours": high_hours,
        "low_tech": low_tech,
        "pain": pain,
        "willing": willing,
        "external": external,
        "learning": learning,
        "avg_hours": sum(row["horas_estimadas"] for row in rows) / total if total else 0,
        "avg_opportunity": sum(row["oportunidad"] for row in rows) / total if total else 0,
        "segments": Counter(row["segmento"] for row in rows),
        "levels": Counter(row["nivel"] for row in rows),
        "processes": processes,
        "goals": goals,
        "barriers": barriers,
        "top_processes": processes.most_common(8),
        "top_goals": goals.most_common(8),
        "top_barriers": barriers.most_common(5),
    }


def fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def border(color: str = LINE) -> Border:
    side = Side(style="thin", color=color)
    return Border(left=side, right=side, top=side, bottom=side)


def set_title(ws, title: str, subtitle: str, end_column: int = 14) -> None:
    end = get_column_letter(end_column)
    ws.merge_cells(f"A1:{end}1")
    ws["A1"] = title
    ws["A1"].font = Font(name="Aptos Display", size=24, bold=True, color=WHITE)
    ws["A1"].fill = fill(NAVY)
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 42
    ws.merge_cells(f"A2:{end}2")
    ws["A2"] = subtitle
    ws["A2"].font = Font(name="Aptos", size=11, color=MUTED, italic=True)
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[2].height = 28
    ws.sheet_view.showGridLines = False


def style_header(row, color: str = NAVY) -> None:
    for cell in row:
        cell.fill = fill(color)
        cell.font = Font(name="Aptos", size=10, bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = border(color)


def style_body(ws, min_row: int, max_row: int, max_col: int) -> None:
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.font = Font(name="Aptos", size=10, color=INK)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=Side(style="hair", color=LINE))
            if cell.row % 2 == 0:
                cell.fill = fill("F7FAFD")


def set_widths(ws, widths: dict[str, float]) -> None:
    for column, width in widths.items():
        ws.column_dimensions[column].width = width


def add_table(ws, name: str, ref: str, style_name: str = "TableStyleMedium2") -> None:
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(name=style_name, showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    ws.add_table(table)


def build_cover(wb: Workbook, stats: dict) -> None:
    ws = wb.active
    ws.title = "00_Portada"
    set_title(ws, "ONDIGITAL · Inteligencia de oportunidad", "Paquete ejecutivo de análisis de encuesta · generado el " + date.today().isoformat(), 8)
    ws.merge_cells("A4:H4")
    ws["A4"] = "La lectura en una frase"
    ws["A4"].font = Font(size=12, bold=True, color=BLUE)
    ws["A5"] = "Ordenar primero. Automatizar después. IA al final."
    ws.merge_cells("A5:H6")
    ws["A5"].font = Font(name="Aptos Display", size=22, bold=True, color=INK)
    ws["A5"].alignment = Alignment(wrap_text=True, vertical="center")
    ws["A5"].fill = fill(PALE_BLUE)
    ws["A5"].border = border("B9D5F7")

    metadata = [
        ("Base analizada", f"{stats['total']} respuestas"),
        ("Periodo visible", "14 de agosto de 2026"),
        ("Tipo de lectura", "Descriptiva y exploratoria"),
        ("Uso recomendado", "Feria, defensa y diagnóstico comercial"),
        ("Fuente", "Encuesta ONDIGITAL-MicroEmpresa.csv"),
    ]
    ws["A9"] = "Ficha del informe"
    ws["A9"].font = Font(size=12, bold=True, color=BLUE)
    for row_index, (label, value) in enumerate(metadata, 10):
        ws.cell(row_index, 1, label)
        ws.cell(row_index, 2, value)
        ws.cell(row_index, 1).font = Font(bold=True, color=MUTED)
        ws.cell(row_index, 2).font = Font(color=INK)
        ws.cell(row_index, 1).border = border()
        ws.cell(row_index, 2).border = border()
        ws.cell(row_index, 2).alignment = Alignment(wrap_text=True)

    ws["E9"] = "Señales para recordar"
    ws["E9"].font = Font(size=12, bold=True, color=BLUE)
    takeaways = [
        (f"{pct(stats['manual'], stats['total']):.1%}", "usa papel y/o Excel"),
        (f"{pct(stats['willing'], stats['total']):.1%}", "está dispuesto a un sistema a la medida"),
        (f"{pct(stats['external'], stats['total']):.1%}", "quiere acompañamiento tecnológico"),
        (f"{pct(stats['barriers'].most_common(1)[0][1], stats['total']):.1%}", "se frena por desconocimiento de opciones"),
    ]
    for row_index, (number, label) in enumerate(takeaways, 10):
        ws.cell(row_index, 5, number)
        ws.cell(row_index, 6, label)
        ws.merge_cells(start_row=row_index, start_column=6, end_row=row_index, end_column=8)
        ws.cell(row_index, 5).font = Font(size=16, bold=True, color=BLUE)
        ws.cell(row_index, 6).font = Font(size=10, color=INK)
        ws.cell(row_index, 5).fill = fill(PALE_MINT)
        ws.cell(row_index, 6).fill = fill(PALE_MINT)
        ws.cell(row_index, 5).border = border("B7E8DA")
        ws.cell(row_index, 6).border = border("B7E8DA")
        ws.cell(row_index, 6).alignment = Alignment(wrap_text=True, vertical="center")
        ws.row_dimensions[row_index].height = 28

    ws.merge_cells("A17:H17")
    ws["A17"] = "Advertencia metodológica"
    ws["A17"].font = Font(size=12, bold=True, color="8A5A00")
    ws["A18"] = "La base muestra señales de respuestas simuladas o altamente guiadas: las 321 respuestas abiertas siguen la misma estructura y 176 pares consecutivos aparecen en el mismo segundo. Presenta estos porcentajes como hallazgos de esta base, no como representatividad del mercado hondureño."
    ws.merge_cells("A18:H20")
    ws["A18"].fill = fill(PALE_AMBER)
    ws["A18"].font = Font(size=10, color="6D4A00")
    ws["A18"].alignment = Alignment(wrap_text=True, vertical="center")
    ws["A18"].border = border("F1D38A")

    ws["A23"] = "Navegación del paquete"
    ws["A23"].font = Font(size=12, bold=True, color=BLUE)
    navigation = [
        ("01_Dashboard", "Indicadores, gráficos y lectura ejecutiva"),
        ("02_Decisiones", "Qué significa y qué haría ONDIGITAL"),
        ("03_Respuestas", "Base limpia y enriquecida, filtrable"),
        ("04_Frecuencias", "Distribuciones completas por pregunta"),
        ("05_Cruces", "Asociaciones y calidad del dato"),
        ("06_Diccionario", "Definiciones, scores y límites"),
        ("07_Visuales", "Gráficos listos para presentar"),
    ]
    for row_index, (sheet, description) in enumerate(navigation, 24):
        ws.cell(row_index, 1, sheet)
        ws.cell(row_index, 2, description)
        ws.cell(row_index, 1).font = Font(bold=True, color=BLUE)
        ws.cell(row_index, 2).font = Font(color=MUTED)
        ws.cell(row_index, 1).hyperlink = f"#'{sheet}'!A1"
        ws.cell(row_index, 1).style = "Hyperlink"
    set_widths(ws, {"A": 25, "B": 31, "C": 3, "D": 3, "E": 18, "F": 18, "G": 18, "H": 18})
    ws.freeze_panes = "A9"
    ws.sheet_properties.tabColor = MINT
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_area = "A1:H30"


def build_dashboard(wb: Workbook, stats: dict, metrics: dict) -> None:
    ws = wb.create_sheet("01_Dashboard")
    set_title(ws, "ONDIGITAL · Dashboard de oportunidad", "Una sala de decisión: señales de dolor, apertura, encaje y calidad de evidencia.", 14)
    ws.merge_cells("A4:N4")
    ws["A4"] = "Veredicto ejecutivo: existe una oportunidad plausible, pero la entrada correcta es ordenar una operación concreta y demostrar valor antes de vender IA."
    ws["A4"].font = Font(size=12, bold=True, color=INK)
    ws["A4"].fill = fill(PALE_BLUE)
    ws["A4"].alignment = Alignment(wrap_text=True, vertical="center")
    ws["A4"].border = border("B9D5F7")
    ws.row_dimensions[4].height = 32

    ws["A6"] = "Indicadores de la muestra"
    ws["A6"].font = Font(size=12, bold=True, color=BLUE)
    metric_rows = [
        ("Tecnología baja o muy baja", stats["low_tech"]),
        ("Pierde 5+ h/semana", stats["high_hours"]),
        ("Usa papel y/o Excel", stats["manual"]),
        ("Percibe pérdida 4–5/5", stats["pain"]),
        ("Dispuesto o muy dispuesto", stats["willing"]),
        ("Quiere equipo externo", stats["external"]),
        ("Dispuesto a aprender", stats["learning"]),
    ]
    headers = ["Indicador", "n", "% de la muestra"]
    for column, value in enumerate(headers, 1):
        ws.cell(7, column, value)
    style_header(ws[7])
    for row_index, (label, count_value) in enumerate(metric_rows, 8):
        ws.cell(row_index, 1, label)
        ws.cell(row_index, 2, count_value)
        ws.cell(row_index, 3, pct(count_value, stats["total"]))
        ws.cell(row_index, 3).number_format = "0.0%"
    style_body(ws, 8, 14, 3)
    add_table(ws, "TblIndicadores", "A7:C14")

    ws["E6"] = "Segmentación comercial"
    ws["E6"].font = Font(size=12, bold=True, color=BLUE)
    for column, value in enumerate(["Segmento", "n", "%"], 5):
        ws.cell(7, column, value)
    style_header(ws[7][4:7])
    segment_order = ["Listo para vender", "Interes medio / diagnostico", "Sensible al precio", "Educar y nutrir", "Baja prioridad"]
    for row_index, label in enumerate(segment_order, 8):
        count_value = stats["segments"].get(label, 0)
        ws.cell(row_index, 5, label)
        ws.cell(row_index, 6, count_value)
        ws.cell(row_index, 7, pct(count_value, stats["total"]))
        ws.cell(row_index, 7).number_format = "0.0%"
    style_body(ws, 8, 12, 7)
    add_table(ws, "TblSegmentos", "E7:G12")

    ws["I6"] = "Calidad y límites"
    ws["I6"].font = Font(size=12, bold=True, color=BLUE)
    quality = [
        ("Respuestas abiertas con plantilla", metrics["calidad_datos"]["q10_parseable_misma_plantilla"]),
        ("Pares en el mismo segundo", metrics["calidad_datos"]["intervalos_consecutivos_mismo_segundo"]),
        ("Pares separados ≤ 2 segundos", metrics["calidad_datos"]["intervalos_consecutivos_hasta_2_segundos"]),
        ("Filas en patrones cerrados", metrics["calidad_datos"]["filas_en_patrones_cerrados_duplicados"]),
        ("Índice de oportunidad promedio", metrics["indice_oportunidad_promedio"]),
    ]
    for row_index, (label, value) in enumerate(quality, 8):
        ws.cell(row_index, 9, label)
        ws.cell(row_index, 10, value)
        ws.merge_cells(start_row=row_index, start_column=10, end_row=row_index, end_column=11)
        ws.cell(row_index, 9).font = Font(size=9, color=MUTED)
        ws.cell(row_index, 10).font = Font(size=11, bold=True, color=INK)
        ws.cell(row_index, 9).border = border()
        ws.cell(row_index, 10).border = border()
        ws.cell(row_index, 10).alignment = Alignment(wrap_text=True)
    ws["J12"].number_format = "0.0"

    chart = BarChart()
    chart.type = "bar"
    chart.style = 10
    chart.title = "Señales de oportunidad"
    chart.y_axis.title = "Indicador"
    chart.x_axis.title = "%"
    chart.height = 8
    chart.width = 14
    chart.legend = None
    chart.add_data(Reference(ws, min_col=3, min_row=7, max_row=14), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=8, max_row=14))
    chart.x_axis.numFmt = "0%"
    ws.add_chart(chart, "A17")

    donut = DoughnutChart()
    donut.title = "Segmentos"
    donut.style = 26
    donut.holeSize = 58
    donut.height = 8
    donut.width = 12
    donut.add_data(Reference(ws, min_col=6, min_row=7, max_row=12), titles_from_data=True)
    donut.set_categories(Reference(ws, min_col=5, min_row=8, max_row=12))
    donut.dataLabels = DataLabelList()
    donut.dataLabels.showPercent = True
    donut.dataLabels.showLeaderLines = True
    ws.add_chart(donut, "I17")

    ws["A34"] = "Prioridades de construcción"
    ws["A34"].font = Font(size=12, bold=True, color=BLUE)
    priorities = [
        ("01", "Centralización", "Un solo lugar para información importante, responsables y estados."),
        ("02", "Workflow", "Tareas recurrentes, alertas, seguimiento y fechas visibles."),
        ("03", "Reportes", "Indicadores simples para decidir sin revisar varios archivos."),
        ("04", "Atención", "Consultas, clientes y respuesta rápida desde un flujo trazable."),
        ("05", "Vito después", "Asistente encima de datos ordenados, no como primera promesa."),
    ]
    for column, value in enumerate(["Orden", "Módulo", "Qué demostrar"], 1):
        ws.cell(35, column, value)
    style_header(ws[35][:3], BLUE)
    for row_index, priority in enumerate(priorities, 36):
        for column, value in enumerate(priority, 1):
            ws.cell(row_index, column, value)
    style_body(ws, 36, 40, 3)
    add_table(ws, "TblPrioridades", "A35:C40")
    ws["E34"] = "Lectura comercial"
    ws["E34"].font = Font(size=12, bold=True, color=BLUE)
    ws.merge_cells("E35:N40")
    ws["E35"] = "El principal freno es el desconocimiento de opciones, no el costo. La conversación de venta debe empezar con un diagnóstico visible del proceso, una demostración breve y un camino modular. Business encaja con quienes quieren continuidad; Vito debe aparecer cuando la operación ya produce datos confiables."
    ws["E35"].fill = fill(PALE_MINT)
    ws["E35"].font = Font(size=11, color=INK)
    ws["E35"].alignment = Alignment(wrap_text=True, vertical="center")
    ws["E35"].border = border("B7E8DA")

    set_widths(ws, {"A": 31, "B": 11, "C": 15, "D": 3, "E": 28, "F": 10, "G": 12, "H": 3, "I": 33, "J": 15, "K": 15, "L": 12, "M": 12, "N": 12})
    ws.freeze_panes = "A7"
    ws.sheet_properties.tabColor = BLUE
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_area = "A1:N40"


def build_decisions(wb: Workbook, stats: dict) -> None:
    ws = wb.create_sheet("02_Decisiones")
    set_title(ws, "ONDIGITAL · Del hallazgo a la acción", "Una traducción operativa para producto, ventas y demostración.", 10)
    ws.merge_cells("A4:J4")
    ws["A4"] = "La encuesta no dice “construyan cualquier software”. Dice: entren por un dolor visible, reduzcan fricción y acompañen el cambio."
    ws["A4"].font = Font(size=12, bold=True, color=INK)
    ws["A4"].fill = fill(PALE_BLUE)
    ws["A4"].alignment = Alignment(wrap_text=True, vertical="center")
    ws["A4"].border = border("B9D5F7")
    ws.row_dimensions[4].height = 32
    headers = ["Hallazgo", "Evidencia en la muestra", "Qué significa para ONDIGITAL", "Siguiente movimiento"]
    for column, value in enumerate(headers, 1):
        ws.cell(6, column, value)
    style_header(ws[6], BLUE)
    barrier_label, barrier_count = stats["top_barriers"][0]
    rows = [
        ("La operación sigue fragmentada", f"{pct(stats['manual'], stats['total']):.1%} usa papel y/o Excel.", "La primera promesa debe ser visibilidad y orden.", "Demo de un flujo completo: entrada → responsable → estado → reporte."),
        ("Hay tiempo perdido suficiente", f"{pct(stats['high_hours'], stats['total']):.1%} pierde 5+ horas semanales.", "El ROI debe expresarse en horas recuperadas, no en jerga técnica.", "Simulador de horas + prueba de un proceso prioritario."),
        ("Existe apertura a personalizar", f"{pct(stats['willing'], stats['total']):.1%} está dispuesto o muy dispuesto.", "La propuesta modular tiene espacio si se demuestra con claridad.", "Ofrecer diagnóstico y módulo inicial, no transformación total."),
        ("El freno principal es pedagógico", f"{barrier_label}: {pct(barrier_count, stats['total']):.1%}.", "La venta debe educar: qué se puede resolver y cómo se implementa.", "Antes/después visual y acompañamiento continuo en lenguaje simple."),
        ("El soporte continuo es parte del valor", f"{pct(stats['external'], stats['total']):.1%} quiere equipo externo.", "Business no es solo hosting: es continuidad y evolución.", "Mostrar mantenimiento, mejoras, monitoreo y contacto humano."),
        ("Vito depende de datos ordenados", f"Índice de oportunidad promedio: {stats['avg_opportunity']:.1f}/100.", "IA no es el primer módulo; es la capa que amplifica una operación consistente.", "Presentar Vito sobre inventario, ventas, agenda o reportes ya centralizados."),
    ]
    for row_index, row in enumerate(rows, 7):
        for column, value in enumerate(row, 1):
            ws.cell(row_index, column, value)
    style_body(ws, 7, 12, 4)
    add_table(ws, "TblDecisiones", "A6:D12", "TableStyleMedium4")
    ws["A15"] = "Guion de demostración recomendado"
    ws["A15"].font = Font(size=12, bold=True, color=BLUE)
    script = [
        ("01 · Mostrar el problema", "Un proceso repetitivo, disperso o sin responsables."),
        ("02 · Ordenar el flujo", "Una pantalla para capturar, asignar, seguir y consultar."),
        ("03 · Medir el cambio", "Horas, estado, pendientes y reportes que antes costaban trabajo."),
        ("04 · Abrir la puerta a Vito", "Preguntas y acciones sobre datos que ya tienen estructura."),
    ]
    for row_index, (step, copy) in enumerate(script, 16):
        ws.cell(row_index, 1, step)
        ws.cell(row_index, 2, copy)
        ws.merge_cells(start_row=row_index, start_column=2, end_row=row_index, end_column=4)
        ws.cell(row_index, 1).font = Font(bold=True, color=BLUE)
        ws.cell(row_index, 2).font = Font(color=INK)
        ws.cell(row_index, 1).fill = fill(PALE_MINT)
        ws.cell(row_index, 2).fill = fill(PALE_MINT)
        ws.cell(row_index, 1).border = border("B7E8DA")
        ws.cell(row_index, 2).border = border("B7E8DA")
        ws.cell(row_index, 2).alignment = Alignment(wrap_text=True)
    ws["A22"] = "Límite de interpretación"
    ws["A22"].font = Font(size=12, bold=True, color="8A5A00")
    ws.merge_cells("A23:D25")
    ws["A23"] = "La muestra se presenta como exploratoria. No usar estos porcentajes para afirmar que representan a todas las microempresas hondureñas. La siguiente fase debe ser una recolección real, con muestra documentada y control de duplicados."
    ws["A23"].fill = fill(PALE_AMBER)
    ws["A23"].font = Font(color="6D4A00")
    ws["A23"].alignment = Alignment(wrap_text=True, vertical="center")
    ws["A23"].border = border("F1D38A")
    set_widths(ws, {"A": 28, "B": 28, "C": 42, "D": 42, "E": 4, "F": 4, "G": 4, "H": 4, "I": 4, "J": 4})
    for row_index in range(7, 13):
        ws.row_dimensions[row_index].height = 56
    ws.freeze_panes = "A7"
    ws.sheet_properties.tabColor = MINT
    ws.page_setup.orientation = "landscape"
    ws.print_area = "A1:D25"


def build_responses(wb: Workbook, rows: list[dict]) -> None:
    ws = wb.create_sheet("03_Respuestas")
    title = "Base limpia y enriquecida"
    set_title(ws, title, "321 registros · filtros, tabla, scores e índice de oportunidad. Los textos abiertos se conservan completos.", 24)
    headers = [
        "ID", "Timestamp", "Madurez tecnológica", "Tiempo perdido", "Almacenamiento actual", "Dolor percibido (1–5)",
        "Disposición a sistema", "Freno principal", "Equipo externo", "Disposición a aprender", "Factores de decisión",
        "Respuesta abierta", "Proceso prioritario", "Problema operativo", "Objetivo", "Score madurez (1–4)",
        "Horas estimadas / semana", "Score dolor (1–5)", "Score disposición (1–4)", "Score externo (1–4)",
        "Score aprendizaje (1–4)", "Índice oportunidad (0–100)", "Nivel oportunidad", "Segmento comercial",
    ]
    for column, value in enumerate(headers, 1):
        ws.cell(4, column, value)
    style_header(ws[4], NAVY)
    for row_index, row in enumerate(rows, 5):
        values = [
            row["id"], row["timestamp"], row["madurez"], row["horas"], "; ".join(row["almacenamiento"]), row["dolor"],
            row["disposicion"], row["freno"], row["externo"], row["aprendizaje"], "; ".join(row["factores"]), row["respuesta"],
            row["proceso"], row["problema"], row["objetivo"], row["score_madurez"], row["horas_estimadas"], row["score_dolor"],
            row["score_disposicion"], row["score_externo"], row["score_aprendizaje"], row["oportunidad"], row["nivel"], row["segmento"],
        ]
        for column, value in enumerate(values, 1):
            ws.cell(row_index, column, value)
    style_body(ws, 5, 4 + len(rows), len(headers))
    add_table(ws, "TblRespuestas", f"A4:X{4 + len(rows)}", "TableStyleMedium2")
    ws.freeze_panes = "C5"
    ws.auto_filter.ref = f"A4:X{4 + len(rows)}"
    ws.sheet_view.zoomScale = 80
    ws.conditional_formatting.add(f"V5:V{4 + len(rows)}", ColorScaleRule(start_type="min", start_color=CORAL, mid_type="percentile", mid_value=50, mid_color=AMBER, end_type="max", end_color=GREEN))
    ws.conditional_formatting.add(f"X5:X{4 + len(rows)}", FormulaRule(formula=['$X5="Listo para vender"'], fill=fill(PALE_MINT)))
    ws.conditional_formatting.add(f"X5:X{4 + len(rows)}", FormulaRule(formula=['$X5="Baja prioridad"'], fill=fill(PALE_CORAL)))
    set_widths(ws, {"A": 7, "B": 25, "C": 16, "D": 24, "E": 34, "F": 14, "G": 18, "H": 33, "I": 17, "J": 19, "K": 35, "L": 56, "M": 24, "N": 25, "O": 27, "P": 14, "Q": 17, "R": 14, "S": 17, "T": 14, "U": 19, "V": 16, "W": 17, "X": 25})
    for row_index in range(5, min(20, 5 + len(rows))):
        ws.row_dimensions[row_index].height = 40
    ws.sheet_properties.tabColor = "7FA7D8"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True


def write_csv_sheet(ws, title: str, rows: list[list[str]], tab_color: str) -> None:
    set_title(ws, title, "Distribuciones completas exportadas desde el análisis reproducible.", 8)
    start_row = 4
    for row_index, values in enumerate(rows, start_row):
        for column, value in enumerate(values, 1):
            ws.cell(row_index, column, value)
    if rows:
        style_header(ws[start_row], NAVY)
        style_body(ws, start_row + 1, start_row + len(rows) - 1, max(len(row) for row in rows))
        last_col = get_column_letter(max(len(row) for row in rows))
        table_name = re.sub(r"[^A-Za-z0-9_]", "", title.replace("_", ""))[:15] or "Datos"
        add_table(ws, f"Tbl{table_name}", f"A{start_row}:{last_col}{start_row + len(rows) - 1}")
    ws.freeze_panes = f"A{start_row + 1}"
    ws.sheet_properties.tabColor = tab_color
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    set_widths(ws, {"A": 14, "B": 62, "C": 15, "D": 16, "E": 16, "F": 16, "G": 16, "H": 16})


def build_frequencies(wb: Workbook) -> None:
    path = ROOT / "tablas" / "frecuencias_completas.csv"
    with path.open(encoding="utf-8-sig", newline="") as source:
        rows = list(csv.reader(source))
    write_csv_sheet(wb.create_sheet("04_Frecuencias"), "Frecuencias completas", rows, AMBER)


def build_crosses(wb: Workbook, metrics: dict) -> None:
    ws = wb.create_sheet("05_Cruces")
    set_title(ws, "Cruces, asociaciones y calidad", "Relaciones descriptivas para orientar hipótesis; no prueban causalidad ni representatividad.", 10)
    ws["A4"] = "Asociaciones principales"
    ws["A4"].font = Font(size=12, bold=True, color=BLUE)
    assoc_path = ROOT / "tablas" / "asociaciones_estadisticas.csv"
    with assoc_path.open(encoding="utf-8-sig", newline="") as source:
        assoc = list(csv.reader(source))
    for row_index, values in enumerate(assoc, 6):
        for column, value in enumerate(values, 1):
            ws.cell(row_index, column, value)
    style_header(ws[6], NAVY)
    style_body(ws, 7, 5 + len(assoc) - 1, 3)
    add_table(ws, "TblAsociaciones", f"A6:C{5 + len(assoc)}")
    ws.conditional_formatting.add(f"B7:B{5 + len(assoc)}", ColorScaleRule(start_type="min", start_color=PALE_CORAL, mid_type="percentile", mid_value=50, mid_color=PALE_AMBER, end_type="max", end_color=PALE_MINT))

    ws["E4"] = "Control de calidad"
    ws["E4"].font = Font(size=12, bold=True, color=BLUE)
    quality = [
        ("Respuestas abiertas con plantilla", metrics["calidad_datos"]["q10_parseable_misma_plantilla"]),
        ("Pares en el mismo segundo", metrics["calidad_datos"]["intervalos_consecutivos_mismo_segundo"]),
        ("Pares ≤ 2 segundos", metrics["calidad_datos"]["intervalos_consecutivos_hasta_2_segundos"]),
        ("Filas en patrones cerrados", metrics["calidad_datos"]["filas_en_patrones_cerrados_duplicados"]),
        ("Duración de captura (s)", metrics["calidad_datos"]["duracion_captura_segundos"]),
    ]
    for row_index, (label, value) in enumerate(quality, 6):
        ws.cell(row_index, 5, label)
        ws.cell(row_index, 6, value)
        ws.cell(row_index, 5).font = Font(color=MUTED, size=10)
        ws.cell(row_index, 6).font = Font(color=INK, bold=True)
        ws.cell(row_index, 5).border = border()
        ws.cell(row_index, 6).border = border()
    ws["E13"] = "Lectura"
    ws["E13"].font = Font(size=12, bold=True, color=BLUE)
    ws.merge_cells("E14:J18")
    ws["E14"] = "La asociación más fuerte del archivo es equipo externo vs aprendizaje (rho = 0.638). También aparecen relaciones moderadas entre tiempo perdido, madurez, frenos y disposición. Úsalas para priorizar preguntas de validación, no como evidencia causal."
    ws["E14"].fill = fill(PALE_BLUE)
    ws["E14"].alignment = Alignment(wrap_text=True, vertical="center")
    ws["E14"].border = border("B9D5F7")
    set_widths(ws, {"A": 42, "B": 18, "C": 23, "D": 4, "E": 34, "F": 18, "G": 18, "H": 18, "I": 18, "J": 18})
    ws.freeze_panes = "A7"
    ws.sheet_properties.tabColor = CORAL
    ws.page_setup.orientation = "landscape"


def build_dictionary(wb: Workbook) -> None:
    ws = wb.create_sheet("06_Diccionario")
    set_title(ws, "Diccionario y metodología", "Definiciones operativas para leer el archivo sin perder contexto.", 8)
    rows = [
        ["Campo", "Descripción / regla", "Tipo", "Nota de uso"],
        ["Índice oportunidad", "Combinación heurística 0–100 de necesidad, disposición, equipo externo, aprendizaje y señales de operación manual.", "Score", "No es probabilidad de compra."],
        ["Necesidad digital", "Resume madurez baja, horas perdidas, dolor percibido y trabajo manual.", "Score", "Sirve para ordenar conversación."],
        ["Encaje ONDIGITAL", "Combina necesidad con apertura a sistema a la medida, equipo externo y aprendizaje.", "Score", "Debe validarse con precio y sector."],
        ["Listo para vender", "Segmento heurístico con necesidad, disposición y condiciones favorables.", "Segmento", "No sustituye un lead calificado."],
        ["Educar y nutrir", "Apertura o necesidad con una objeción que requiere demostración y acompañamiento.", "Segmento", "Siguiente paso: diagnóstico."],
        ["Sensible al precio", "Hay interés, pero el costo aparece como freno dominante o la propuesta necesita un caso de valor.", "Segmento", "Siguiente paso: ROI y entrada modular."],
        ["Interés medio / diagnóstico", "Señales mixtas; conviene entender el proceso antes de ofertar.", "Segmento", "No perseguir con mensaje genérico."],
        ["Baja prioridad", "Pocas señales de dolor o disposición inmediata.", "Segmento", "No forzar la venta."],
        ["Horas estimadas", "Punto medio de la categoría de horas declarada: 0, 2.5, 7.5 o 12.5.", "Derivado", "No es una medición de tiempo real."],
        ["Pregunta abierta", "Se conserva completa y se clasifica en proceso, problema y objetivo mediante una plantilla gramatical.", "Texto", "La repetición exacta es una alerta metodológica."],
        ["Base de referencia", "321 registros capturados el 14 de agosto de 2026.", "Contexto", "La evidencia describe este archivo."],
        ["Advertencia", "La base parece simulada, generada o fuertemente guiada por estructura temporal y textual.", "Calidad", "No extrapolar a todo Honduras."],
    ]
    for row_index, values in enumerate(rows, 4):
        for column, value in enumerate(values, 1):
            ws.cell(row_index, column, value)
    style_header(ws[4], NAVY)
    style_body(ws, 5, 3 + len(rows) - 1, 4)
    add_table(ws, "TblDiccionario", f"A4:D{3 + len(rows)}", "TableStyleMedium4")
    set_widths(ws, {"A": 25, "B": 75, "C": 16, "D": 42, "E": 4, "F": 4, "G": 4, "H": 4})
    for row_index in range(5, 4 + len(rows)):
        ws.row_dimensions[row_index].height = 42
    ws.freeze_panes = "A5"
    ws.sheet_properties.tabColor = AMBER
    ws.page_setup.orientation = "landscape"


def build_visuals(wb: Workbook) -> None:
    ws = wb.create_sheet("07_Visuales")
    set_title(ws, "ONDIGITAL · Visuales listos para presentar", "Selección de gráficos del análisis original, organizada para revisión y exposición.", 18)
    if XLImage is None:
        ws["A4"] = "Instala openpyxl con soporte de imágenes para mostrar los gráficos en esta hoja."
        return
    images = [
        ("Madurez tecnológica", ROOT / "graficos" / "01_madurez_tecnologica.png", "A4"),
        ("Disposición a sistema", ROOT / "graficos" / "05_disposicion_sistema_medida.png", "J4"),
        ("Segmentos comerciales", ROOT / "graficos" / "14_segmentos_comerciales.png", "A23"),
        ("Índice de oportunidad", ROOT / "graficos" / "15_indice_oportunidad.png", "J23"),
    ]
    for label, path, anchor in images:
        cell = ws[anchor]
        cell.value = label
        cell.font = Font(size=12, bold=True, color=BLUE)
        image = XLImage(path)
        original_width, original_height = image.width, image.height
        image.width = 520
        image.height = int(520 * original_height / original_width)
        ws.add_image(image, anchor.replace(str(cell.row), str(cell.row + 1)))
    set_widths(ws, {"A": 14, "B": 14, "C": 14, "D": 14, "E": 14, "F": 14, "G": 4, "H": 4, "I": 4, "J": 14, "K": 14, "L": 14, "M": 14, "N": 14, "O": 14, "P": 4, "Q": 4, "R": 4})
    ws.sheet_properties.tabColor = "7FA7D8"
    ws.page_setup.orientation = "landscape"
    ws.sheet_view.showGridLines = False


def build_workbook(rows: list[dict], metrics: dict) -> None:
    stats = summary(rows)
    wb = Workbook()
    build_cover(wb, stats)
    build_dashboard(wb, stats, metrics)
    build_decisions(wb, stats)
    build_responses(wb, rows)
    build_frequencies(wb)
    build_crosses(wb, metrics)
    build_dictionary(wb)
    build_visuals(wb)
    wb.properties.title = "ONDIGITAL · Análisis de oportunidad digital"
    wb.properties.subject = "Encuesta ONDIGITAL-MicroEmpresa"
    wb.properties.creator = "ONDIGITAL"
    wb.properties.description = "Paquete ejecutivo y exploratorio; resultados descriptivos de una base de 321 respuestas."
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    wb.save(OUTPUT_XLSX)


EXPLORER_TEMPLATE = r'''<!doctype html>
<html lang="es">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<meta name="color-scheme" content="dark light">
<title>ONDIGITAL | Explorador de respuestas</title>
<style>
:root{--bg:#071426;--panel:#0d1d33;--panel-2:#102542;--ink:#eef7ff;--muted:#a9bad1;--line:rgba(177,210,244,.18);--mint:#00e5b0;--blue:#3b82f6;--green:#49d18e;--amber:#f9b84a;--coral:#ff6b75;--shadow:0 18px 48px rgba(0,0,0,.23);--radius:16px;--font:Inter,ui-sans-serif,system-ui,-apple-system,"Segoe UI",sans-serif;--mono:"SFMono-Regular",Consolas,monospace}
body.light{--bg:#f4f7fb;--panel:#fff;--panel-2:#eef3f8;--ink:#0b1a2e;--muted:#5d6d81;--line:#d5dfeb;--mint:#008f78;--blue:#246bd1;--green:#167a4a;--amber:#a66500;--coral:#b9362d;--shadow:0 14px 38px rgba(22,48,81,.12)}
*{box-sizing:border-box}html{scroll-behavior:smooth}body{margin:0;background-color:var(--bg);background-image:linear-gradient(rgba(82,144,211,.045) 1px,transparent 1px),linear-gradient(90deg,rgba(82,144,211,.045) 1px,transparent 1px);background-size:32px 32px;color:var(--ink);font:14px/1.5 var(--font)}a{color:inherit}button,input,select{font:inherit}button{cursor:pointer}button:focus-visible,a:focus-visible,input:focus-visible,select:focus-visible{outline:3px solid color-mix(in srgb,var(--mint) 70%,transparent);outline-offset:3px}.shell{width:min(1440px,calc(100% - 32px));margin:auto}.topbar{position:sticky;top:0;z-index:20;background:color-mix(in srgb,var(--bg) 90%,transparent);border-bottom:1px solid var(--line);backdrop-filter:blur(16px)}.topbar-inner{min-height:64px;display:flex;align-items:center;justify-content:space-between;gap:14px}.brand{display:flex;align-items:center;gap:11px;font-weight:850;letter-spacing:.12em;font-size:13px}.mark{display:grid;place-items:center;width:34px;height:34px;border-radius:10px;background:linear-gradient(135deg,var(--blue),var(--mint));color:#061426;font-weight:950;letter-spacing:-.08em}.top-actions{display:flex;gap:8px;align-items:center}.btn{border:1px solid var(--line);background:var(--panel);color:var(--ink);border-radius:10px;padding:9px 12px;font-weight:750;text-decoration:none;transition:transform .18s ease,border-color .18s ease,box-shadow .18s ease}.btn:hover{transform:translateY(-1px);border-color:var(--blue);box-shadow:0 8px 22px rgba(27,111,205,.17)}.btn.primary{background:var(--mint);color:#061426;border-color:var(--mint)}
.page-head{padding:48px 0 28px;display:flex;align-items:end;justify-content:space-between;gap:24px}.kicker{font:700 11px/1 var(--mono);text-transform:uppercase;letter-spacing:.16em;color:var(--mint);margin-bottom:12px}.page-head h1{font-size:clamp(34px,5vw,66px);line-height:.95;letter-spacing:-.05em;margin:0}.page-head p{max-width:640px;color:var(--muted);margin:13px 0 0;font-size:16px}.page-head .head-note{max-width:310px;border-left:1px solid var(--mint);padding-left:14px;color:var(--muted);font-size:12px}.control-panel{display:grid;grid-template-columns:minmax(240px,2fr) repeat(2,minmax(170px,1fr)) auto;gap:10px;align-items:end;padding:16px;border:1px solid color-mix(in srgb,var(--blue) 25%,var(--line));border-radius:var(--radius);background:var(--panel);box-shadow:var(--shadow)}.field{display:flex;flex-direction:column;gap:6px}.field label{font:700 10px/1 var(--mono);letter-spacing:.08em;text-transform:uppercase;color:var(--muted)}.field input,.field select{width:100%;min-width:0;border:1px solid var(--line);border-radius:10px;background:var(--bg);color:var(--ink);padding:10px 11px}.field input:hover,.field select:hover{border-color:var(--blue)}.control-actions{display:flex;gap:8px;flex-wrap:wrap}.result-bar{display:flex;align-items:center;justify-content:space-between;gap:14px;margin:13px 0;color:var(--muted);font-size:12px}.result-bar strong{color:var(--ink);font-size:15px}.result-bar .status{display:inline-flex;align-items:center;gap:8px}.status i{width:8px;height:8px;border-radius:50%;background:var(--mint);box-shadow:0 0 0 4px color-mix(in srgb,var(--mint) 14%,transparent)}.stats{display:grid;grid-template-columns:repeat(4,1fr);gap:10px;margin:16px 0 26px}.stat{border:1px solid var(--line);border-radius:14px;background:var(--panel);padding:15px}.stat b{display:block;font-size:26px;letter-spacing:-.05em;color:var(--mint)}.stat span{display:block;color:var(--muted);font-size:11px;margin-top:4px}.table-card{border:1px solid var(--line);border-radius:var(--radius);background:var(--panel);box-shadow:var(--shadow);overflow:hidden}.table-head{display:flex;align-items:end;justify-content:space-between;gap:16px;padding:20px 20px 15px;border-bottom:1px solid var(--line)}.table-head h2{font-size:20px;letter-spacing:-.02em;margin:0}.table-head p{color:var(--muted);margin:4px 0 0;font-size:12px}.table-tools{display:flex;gap:8px;flex-wrap:wrap}.table-scroll{overflow:auto;max-height:650px}.table-scroll table{border-collapse:collapse;min-width:1320px;width:100%;font-size:12px}.table-scroll th,.table-scroll td{padding:11px 12px;border-bottom:1px solid var(--line);vertical-align:top;text-align:left}.table-scroll th{position:sticky;top:0;z-index:2;background:var(--panel-2);color:var(--muted);font:700 10px/1.2 var(--mono);text-transform:uppercase;letter-spacing:.04em}.table-scroll tbody tr{transition:background .16s ease}.table-scroll tbody tr:hover,.table-scroll tbody tr:focus{background:color-mix(in srgb,var(--blue) 9%,var(--panel));outline:none}.table-scroll td:first-child,.table-scroll th:first-child{position:sticky;left:0;z-index:3;background:var(--panel);font-family:var(--mono);color:var(--muted)}.table-scroll th:first-child{background:var(--panel-2)}.wrap{white-space:normal;min-width:220px;max-width:310px}.score{font-weight:850;color:var(--mint);font-family:var(--mono)}.segment{display:inline-flex;border-radius:999px;padding:5px 8px;background:color-mix(in srgb,var(--blue) 13%,transparent);color:var(--blue);font-size:10px;font-weight:800;white-space:nowrap}.pagination{display:flex;align-items:center;justify-content:space-between;gap:12px;padding:14px 20px;color:var(--muted);font-size:12px}.pagination-actions{display:flex;gap:8px}.btn:disabled{cursor:not-allowed;opacity:.45;transform:none;box-shadow:none}.empty{padding:42px 18px;text-align:center;color:var(--muted)}dialog{width:min(720px,calc(100% - 32px));max-height:90vh;overflow:auto;border:1px solid var(--line);border-radius:16px;background:var(--panel);color:var(--ink);padding:0;box-shadow:0 25px 80px rgba(0,0,0,.4)}dialog::backdrop{background:rgba(1,9,20,.68);backdrop-filter:blur(5px)}.dialog-head{display:flex;justify-content:space-between;gap:12px;padding:18px 20px;border-bottom:1px solid var(--line)}.dialog-head h2{font-size:19px;margin:0}.dialog-head p{color:var(--muted);font:11px var(--mono);margin:5px 0 0}.dialog-body{padding:18px 20px}.detail-grid{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:10px}.detail{border:1px solid var(--line);border-radius:12px;padding:12px;background:var(--panel-2)}.detail label{display:block;color:var(--muted);font:700 10px/1.2 var(--mono);text-transform:uppercase;margin-bottom:5px}.detail strong{font-size:13px}.detail.full{grid-column:1/-1}.footer{padding:34px 0 60px;color:var(--muted);font-size:11px}.footer strong{color:var(--ink)}
@media(max-width:950px){.page-head{display:block}.page-head .head-note{margin-top:18px}.control-panel{grid-template-columns:1fr 1fr}.field:first-child{grid-column:1/-1}.control-actions{grid-column:1/-1}.stats{grid-template-columns:repeat(2,1fr)}}@media(max-width:560px){.shell{width:min(100% - 20px,1440px)}.topbar-inner{min-height:58px}.brand{font-size:11px}.top-actions .text-action{display:none}.page-head{padding:32px 0 21px}.page-head h1{font-size:40px}.page-head p{font-size:14px}.control-panel{grid-template-columns:1fr}.field:first-child,.control-actions{grid-column:auto}.control-actions .btn{flex:1}.result-bar{align-items:flex-start;flex-direction:column}.stats{gap:7px}.stat{padding:12px}.stat b{font-size:22px}.table-head{display:block}.table-tools{margin-top:12px}.table-tools .btn{flex:1}.detail-grid{grid-template-columns:1fr}.detail.full{grid-column:auto}.pagination{align-items:flex-start;flex-direction:column}.pagination-actions{width:100%}.pagination-actions .btn{flex:1}}
@media(prefers-reduced-motion:reduce){*,*::before,*::after{scroll-behavior:auto!important;transition-duration:.01ms!important}}
body{font-family:"Aptos","Inter",ui-sans-serif,system-ui,sans-serif}
.page-head h1{font-family:"Aptos Display","Avenir Next","Inter",ui-sans-serif,sans-serif}

/* Executive data explorer: light-first, dense and easy to scan. */
:root{--bg:#f4f7fa;--panel:#fff;--panel-2:#edf2f7;--ink:#071426;--muted:#5c6d80;--line:#d8e1ea;--mint:#008f78;--blue:#246bd1;--green:#167a4a;--amber:#9a6700;--coral:#b9362d;--navy:#071426;--navy-2:#0d2542;--shadow:0 10px 30px rgba(28,53,82,.08);--radius:10px}
body.dark{--bg:#08111f;--panel:#0d1d2f;--panel-2:#122841;--ink:#edf6ff;--muted:#a7b8ca;--line:rgba(181,207,235,.18);--mint:#28d9b6;--blue:#65a3ff;--green:#5bd49a;--amber:#f2b84b;--coral:#ff7a83;--navy:#040b14;--navy-2:#102b4f;--shadow:0 14px 36px rgba(0,0,0,.22)}
*{letter-spacing:0!important}
body{background:var(--bg);background-image:none;color:var(--ink);font-size:14px}
.shell{width:min(1440px,calc(100% - 40px))}
.topbar{background:color-mix(in srgb,var(--panel) 94%,transparent);border-bottom-color:var(--line);backdrop-filter:blur(14px)}
.topbar-inner{min-height:64px}
.brand{text-decoration:none;font-size:13px}
.brand-copy{display:grid;gap:1px}
.brand-copy strong{font-size:13px;color:var(--ink)}
.brand-copy small{font-size:10px;color:var(--muted);font-weight:700}
.mark{width:32px;height:32px;border-radius:7px;background:var(--navy);color:#eef7ff;border-bottom:3px solid var(--mint);font-size:11px}
.btn{min-height:40px;padding:9px 12px;border-color:var(--line);border-radius:8px;background:var(--panel);color:var(--ink);box-shadow:none;transition:background .18s ease,border-color .18s ease,color .18s ease}
.btn:hover{transform:none;background:var(--panel-2);border-color:var(--blue);box-shadow:none}
.btn.primary{background:var(--navy);color:#eef7ff;border-color:var(--navy)}
.btn.primary:hover{background:var(--navy-2);border-color:var(--navy-2)}
button:focus-visible,a:focus-visible,input:focus-visible,select:focus-visible{outline:3px solid color-mix(in srgb,var(--blue) 42%,transparent);outline-offset:2px}
.page-head{display:grid;grid-template-columns:minmax(0,1fr) minmax(260px,360px);align-items:center;gap:40px;padding:38px 0 24px}
.kicker{margin-bottom:10px;color:var(--mint);font:800 12px/1.2 var(--font);text-transform:none}
.page-head h1{margin:0;color:var(--ink);font-size:44px;line-height:1.05}
.page-head p{max-width:680px;margin:11px 0 0;color:var(--muted);font-size:15px;line-height:1.55}
.page-head .head-note{max-width:none;padding:14px 16px;border:0;border-left:3px solid var(--mint);border-radius:0 8px 8px 0;background:var(--panel);color:var(--muted);font-size:12px}
.control-panel{grid-template-columns:minmax(260px,2fr) repeat(2,minmax(180px,1fr)) auto;gap:12px;padding:16px;border-color:var(--line);border-radius:10px;background:var(--panel);box-shadow:var(--shadow)}
.field{gap:7px}.field label{color:var(--muted);font:750 11px/1.2 var(--font);text-transform:none}
.field input,.field select{min-height:42px;padding:10px 11px;border-color:var(--line);border-radius:8px;background:var(--bg);color:var(--ink)}
.control-actions{gap:8px}
.result-bar{margin:12px 0;color:var(--muted)}
.result-bar strong{color:var(--ink)}
.status i{background:var(--mint);box-shadow:none}
.stats{display:grid;grid-template-columns:repeat(4,1fr);gap:1px;margin:14px 0 24px;padding:1px;border:1px solid var(--line);border-radius:10px;overflow:hidden;background:var(--line)}
.stat{min-height:92px;padding:15px;border:0;border-radius:0;background:var(--panel)}
.stat b{color:var(--mint);font-size:27px}
.stat span{font-size:11px}
.table-card{border-color:var(--line);border-radius:10px;background:var(--panel);box-shadow:0 2px 10px rgba(28,53,82,.05)}
.table-head{align-items:center;padding:17px 18px 14px;border-bottom-color:var(--line)}
.table-head h2{font-size:19px}.table-head p{font-size:12px}
.table-scroll{max-height:680px}
.table-scroll table{font-size:12px}
.table-scroll th,.table-scroll td{padding:11px 12px;border-bottom-color:var(--line)}
.table-scroll th{background:var(--panel-2);color:var(--muted);font-size:10px}
.table-scroll tbody tr:hover,.table-scroll tbody tr:focus{background:color-mix(in srgb,var(--blue) 7%,var(--panel))}
.table-scroll td:first-child{background:var(--panel)}.table-scroll th:first-child{background:var(--panel-2)}
.segment{border-radius:5px;background:color-mix(in srgb,var(--blue) 10%,transparent);color:var(--blue)}
.pagination{padding:13px 18px}
dialog{border-radius:10px;background:var(--panel);color:var(--ink)}
.detail{border-color:var(--line);border-radius:8px;background:var(--panel-2)}
.footer{padding:30px 0 50px}
@media(max-width:950px){.shell{width:min(100% - 24px,1440px)}.page-head{grid-template-columns:1fr;gap:16px}.control-panel{grid-template-columns:1fr 1fr}.stats{grid-template-columns:repeat(2,1fr)}}
@media(max-width:560px){.shell{width:min(100% - 20px,1440px)}.brand{font-size:11px}.brand-copy small{display:none}.top-actions #themeBtn{display:inline-flex}.top-actions .text-action:last-child{display:none}.btn,.field input,.field select{min-height:44px}.page-head{padding:28px 0 20px}.page-head h1{font-size:34px}.page-head p{font-size:14px}.control-panel{grid-template-columns:1fr;padding:14px}.field:first-child,.control-actions{grid-column:auto}.control-actions .btn{flex:1}.stats{grid-template-columns:repeat(2,1fr)}.stat{min-height:84px;padding:12px}.stat b{font-size:23px}.table-head{display:block}.table-tools{margin-top:12px}.table-tools .btn{flex:1}.pagination{align-items:flex-start;flex-direction:column}.pagination-actions{width:100%}.pagination-actions .btn{flex:1}.detail-grid{grid-template-columns:1fr}.detail.full{grid-column:auto}}
@media(prefers-reduced-motion:reduce){*,*::before,*::after{scroll-behavior:auto!important;transition-duration:.01ms!important;animation-duration:.01ms!important}}
</style>
</head>
<body>
<div class="topbar"><div class="shell topbar-inner"><a class="brand" href="dashboard_ondigital.html" aria-label="Volver al dashboard"><span class="mark" aria-hidden="true">OD</span><span class="brand-copy"><strong>ONDIGITAL</strong><small>Datos de encuesta</small></span></a><div class="top-actions"><button class="btn text-action" id="themeBtn" type="button" aria-pressed="false">Oscuro</button><a class="btn text-action" href="dashboard_ondigital.html">Dashboard</a></div></div></div>
<main class="shell">
  <header class="page-head"><div><div class="kicker">Evidencia · 321 respuestas</div><h1>Explorador de respuestas</h1><p>Busca, filtra y abre cada registro sin perder el contexto del análisis. La tabla conserva la respuesta original y sus derivados operativos.</p></div><div class="head-note">La lectura es descriptiva de esta base. No representa por sí sola a todas las microempresas hondureñas.</div></header>
  <section class="control-panel" aria-label="Filtros de respuestas">
    <div class="field"><label for="search">Buscar en toda la respuesta</label><input id="search" type="search" placeholder="inventario, reportes, errores…" autocomplete="off"></div>
    <div class="field"><label for="segment">Segmento</label><select id="segment"><option value="">Todos los segmentos</option><option>Listo para vender</option><option>Educar y nutrir</option><option>Sensible al precio</option><option>Interes medio / diagnostico</option><option>Baja prioridad</option></select></div>
    <div class="field"><label for="level">Nivel de oportunidad</label><select id="level"><option value="">Todos los niveles</option><option>Muy alta</option><option>Alta</option><option>Media</option><option>Baja</option></select></div>
    <div class="control-actions"><button class="btn" id="clearBtn" type="button">Limpiar</button><button class="btn primary" id="downloadBtn" type="button">Descargar CSV</button></div>
  </section>
  <div class="result-bar" aria-live="polite"><span class="status"><i aria-hidden="true"></i><strong id="resultCount">321</strong> respuestas visibles</span><span id="pageHint">Página 1 de 18</span></div>
  <section class="stats" aria-label="Resumen de selección"><div class="stat"><b id="statOpportunity">55.5</b><span>oportunidad promedio / 100</span></div><div class="stat"><b id="statManual">73.2%</b><span>usa papel y/o Excel</span></div><div class="stat"><b id="statPain">42.1%</b><span>dolor percibido 4–5/5</span></div><div class="stat"><b id="statReady">23.4%</b><span>segmento listo para vender</span></div></section>
  <section class="table-card" aria-labelledby="tableTitle"><header class="table-head"><div><h2 id="tableTitle">Registro operativo</h2><p>Haz clic o pulsa Enter sobre una fila para revisar el detalle completo.</p></div><div class="table-tools"><button class="btn" id="prevBtn" type="button">Anterior</button><button class="btn" id="nextBtn" type="button">Siguiente</button></div></header><div class="table-scroll"><table><thead><tr><th>#</th><th>Madurez</th><th>Tiempo</th><th>Dolor</th><th>Disposición</th><th>Freno</th><th>Proceso prioritario</th><th>Problema</th><th>Objetivo</th><th>Oportunidad</th><th>Segmento</th></tr></thead><tbody id="tableBody"></tbody></table></div><div class="pagination"><span id="paginationText">Mostrando 1–18 de 321</span><div class="pagination-actions"><button class="btn" id="prevBottom" type="button">Anterior</button><button class="btn" id="nextBottom" type="button">Siguiente</button></div></div></section>
</main>
<footer class="footer"><div class="shell"><strong>ONDIGITAL · Todo lo vital es digital.</strong><br>Fuente: Encuesta ONDIGITAL-MicroEmpresa. Scores e índices son derivados heurísticos y están documentados en el libro Excel.</div></footer>
<dialog id="detailDialog"><div class="dialog-head"><div><h2 id="dialogTitle">Respuesta</h2><p id="dialogMeta"></p></div><button class="btn" id="closeDialog" type="button">Cerrar</button></div><div class="dialog-body"><div class="detail-grid" id="dialogBody"></div></div></dialog>
<script>
const DATA=__DATA__;
const state={query:'',segment:'',level:'',page:1,pageSize:18,rows:[]};
const $=selector=>document.querySelector(selector), $$=selector=>[...document.querySelectorAll(selector)];
const esc=value=>String(value??'').replace(/[&<>"']/g,match=>({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#039;'}[match]));
const pct=(value,total)=>total?100*value/total:0;
const pctText=value=>`${value.toFixed(1)}%`;
function searchable(row){return [row.madurez,row.horas,row.almacenamiento.join(' '),row.disposicion,row.freno,row.externo,row.aprendizaje,row.factores.join(' '),row.respuesta,row.proceso,row.problema,row.objetivo,row.nivel,row.segmento].join(' ').toLocaleLowerCase('es-HN')}
function getRows(){const query=state.query.toLocaleLowerCase('es-HN');return DATA.filter(row=>(!query||searchable(row).includes(query))&&(!state.segment||row.segmento===state.segment)&&(!state.level||row.nivel===state.level))}
function renderStats(rows){const total=rows.length||1;const avg=rows.reduce((sum,row)=>sum+row.oportunidad,0)/total;const manual=rows.filter(row=>row.almacenamiento.some(item=>/libreta|excel|hoja/i.test(item))).length;const pain=rows.filter(row=>row.dolor>=4).length;const ready=rows.filter(row=>row.segmento==='Listo para vender').length;$('#statOpportunity').textContent=rows.length?avg.toFixed(1):'0.0';$('#statManual').textContent=pctText(pct(manual,total));$('#statPain').textContent=pctText(pct(pain,total));$('#statReady').textContent=pctText(pct(ready,total))}
function shortHours(value){if(value.startsWith('No se'))return '0 h / optimizado';if(value.startsWith('Menos'))return '<5 h';if(value.startsWith('Entre'))return '5–10 h';return '>10 h'}
function shortBarrier(value){if(value.startsWith('Desconocimiento'))return 'Desconocimiento';if(value.startsWith('falta de tiempo'))return 'Falta de tiempo';if(value.startsWith('El costo'))return 'Costo';return value}
function shortSegment(value){return value==='Interes medio / diagnostico'?'Interés medio / diagnóstico':value}
function rowMarkup(row){return `<tr tabindex="0" data-id="${row.id}" aria-label="Respuesta ${row.id}"><td>${row.id}</td><td>${esc(row.madurez)}</td><td>${esc(shortHours(row.horas))}</td><td>${row.dolor}/5</td><td>${esc(row.disposicion)}</td><td class="wrap">${esc(shortBarrier(row.freno))}</td><td class="wrap">${esc(row.proceso)}</td><td class="wrap">${esc(row.problema)}</td><td class="wrap">${esc(row.objetivo)}</td><td class="score">${row.oportunidad.toFixed(1)}</td><td><span class="segment">${esc(shortSegment(row.segmento))}</span></td></tr>`}
function render(){state.rows=getRows();const pages=Math.max(1,Math.ceil(state.rows.length/state.pageSize));state.page=Math.min(state.page,pages);const start=(state.page-1)*state.pageSize;const visible=state.rows.slice(start,start+state.pageSize);$('#tableBody').innerHTML=visible.length?visible.map(rowMarkup).join(''):`<tr><td colspan="11"><div class="empty">No hay respuestas con estos filtros. Prueba una búsqueda más amplia.</div></td></tr>`;$('#resultCount').textContent=state.rows.length;$('#pageHint').textContent=`Página ${state.page} de ${pages}`;$('#paginationText').textContent=state.rows.length?`Mostrando ${start+1}–${Math.min(start+state.pageSize,state.rows.length)} de ${state.rows.length}`:'0 respuestas';$('#prevBtn').disabled=state.page<=1;$('#prevBottom').disabled=state.page<=1;$('#nextBtn').disabled=state.page>=pages;$('#nextBottom').disabled=state.page>=pages;renderStats(state.rows)}
function movePage(delta){const pages=Math.max(1,Math.ceil(state.rows.length/state.pageSize));state.page=Math.max(1,Math.min(pages,state.page+delta));render();document.querySelector('.table-card').scrollIntoView({behavior:'smooth',block:'start'})}
function openDetail(row){$('#dialogTitle').textContent=`Respuesta #${row.id}`;$('#dialogMeta').textContent=row.timestamp;const details=[['Madurez tecnológica',row.madurez],['Tiempo perdido',row.horas],['Almacenamiento',row.almacenamiento.join('; ')],['Dolor percibido',`${row.dolor}/5`],['Disposición a sistema',row.disposicion],['Freno principal',row.freno],['Equipo externo',row.externo],['Aprendizaje',row.aprendizaje],['Proceso prioritario',row.proceso],['Problema operativo',row.problema],['Objetivo',row.objetivo],['Índice oportunidad',row.oportunidad.toFixed(1)],['Nivel',row.nivel],['Segmento',row.segmento],['Factores de decisión',row.factores.join('; ')],['Respuesta original',row.respuesta]];$('#dialogBody').innerHTML=details.map(([label,value],index)=>`<div class="detail ${index>=14?'full':''}"><label>${esc(label)}</label><strong>${esc(value)}</strong></div>`).join('');const dialog=$('#detailDialog');if(dialog.showModal)dialog.showModal();else dialog.setAttribute('open','')}
function downloadCsv(){const columns=['id','timestamp','madurez','horas','almacenamiento','dolor','disposicion','freno','externo','aprendizaje','factores','proceso','problema','objetivo','oportunidad','nivel','segmento'];const lines=[columns.join(','),...state.rows.map(row=>columns.map(key=>{const value=Array.isArray(row[key])?row[key].join('; '):row[key];return `"${String(value??'').replaceAll('"','""')}"`}).join(','))];const blob=new Blob(['\ufeff'+lines.join('\n')],{type:'text/csv;charset=utf-8'});const link=document.createElement('a');link.href=URL.createObjectURL(blob);link.download='ONDIGITAL_respuestas_filtradas.csv';link.click();URL.revokeObjectURL(link.href)}
function syncTheme(){const dark=document.body.classList.contains('dark');const button=$('#themeBtn');button.textContent=dark?'Claro':'Oscuro';button.setAttribute('aria-pressed',String(dark));button.title=dark?'Activar tema claro':'Activar tema oscuro';button.setAttribute('aria-label',button.title)}
$('#search').addEventListener('input',event=>{state.query=event.target.value.trim();state.page=1;render()});$('#segment').addEventListener('change',event=>{state.segment=event.target.value;state.page=1;render()});$('#level').addEventListener('change',event=>{state.level=event.target.value;state.page=1;render()});$('#clearBtn').addEventListener('click',()=>{$('#search').value='';$('#segment').value='';$('#level').value='';state.query='';state.segment='';state.level='';state.page=1;render()});['#prevBtn','#prevBottom'].forEach(selector=>$(selector).addEventListener('click',()=>movePage(-1)));['#nextBtn','#nextBottom'].forEach(selector=>$(selector).addEventListener('click',()=>movePage(1)));$('#downloadBtn').addEventListener('click',downloadCsv);$('#tableBody').addEventListener('click',event=>{const tr=event.target.closest('tr[data-id]');if(tr)openDetail(DATA.find(row=>row.id===Number(tr.dataset.id)))});$('#tableBody').addEventListener('keydown',event=>{if(event.key==='Enter'||event.key===' '){const tr=event.target.closest('tr[data-id]');if(tr){event.preventDefault();openDetail(DATA.find(row=>row.id===Number(tr.dataset.id)))}}});$('#closeDialog').addEventListener('click',()=>$('#detailDialog').close());$('#detailDialog').addEventListener('click',event=>{if(event.target===event.currentTarget)event.currentTarget.close()});$('#themeBtn').addEventListener('click',()=>{document.body.classList.toggle('dark');try{localStorage.setItem('ondigital-explorer-theme',document.body.classList.contains('dark')?'dark':'light')}catch(error){}syncTheme()});try{if(localStorage.getItem('ondigital-explorer-theme')==='dark')document.body.classList.add('dark')}catch(error){};syncTheme();render();
</script>
</body>
</html>
'''


def build_explorer(rows: list[dict]) -> None:
    payload = json.dumps(rows, ensure_ascii=False, separators=(",", ":"))
    OUTPUT_EXPLORER.write_text(EXPLORER_TEMPLATE.replace("__DATA__", payload), encoding="utf-8")


def main() -> None:
    raw_rows, metrics = load_rows()
    rows = enrich_rows(raw_rows)
    build_workbook(rows, metrics)
    build_explorer(rows)
    print(f"Generados: {OUTPUT_XLSX.name} y {OUTPUT_EXPLORER.name} · {len(rows)} respuestas")


if __name__ == "__main__":
    main()
